import csv
import json

from openpyxl import load_workbook

from brightpath import BackgroundProfile, BrightwayInventory, SimaProInventory
from brightpath.analysis import (
    SOURCE_FORMAT_BRIGHTWAY_CSV,
    SOURCE_FORMAT_BRIGHTWAY_EXCEL,
    SOURCE_FORMAT_BRIGHTWAY_TSV,
    SOURCE_FORMAT_SIMAPRO_CSV,
    InventoryValidationError,
    analyze_inventory,
    infer_source_format,
    validate_inventory,
)


def minimal_activity(extra_exchanges=None, **overrides):
    activity = {
        "name": "test process",
        "reference product": "test product",
        "location": "GLO",
        "unit": "kilogram",
        "exchanges": [
            {
                "type": "production",
                "name": "test process",
                "reference product": "test product",
                "location": "GLO",
                "unit": "kilogram",
                "amount": 1.0,
                "simapro category": "Materials/Test",
            }
        ],
    }
    activity.update(overrides)
    if extra_exchanges:
        activity["exchanges"].extend(extra_exchanges)
    return activity


def make_brightway_excel(tmp_path, data, db_name="analysis_db"):
    return BrightwayInventory.from_data(
        data,
        background_profile=BackgroundProfile("ecoinvent", "3.9", "cutoff"),
        database_name=db_name,
    ).write_excel(tmp_path / "inventory", validate=False)


def make_simapro_csv(tmp_path, data, filename="inventory.csv"):
    return SimaProInventory.from_data(
        data,
        background_profile=BackgroundProfile("ecoinvent", "3.9", "cutoff"),
    ).write_csv(tmp_path / filename, validate=False)


def make_brightway_delimited(tmp_path, data, *, delimiter=",", suffix=".csv", db_name="analysis_db"):
    workbook = make_brightway_excel(tmp_path, data, db_name=db_name)
    sheet = load_workbook(workbook, read_only=True, data_only=False).active
    path = tmp_path / f"inventory{suffix}"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])
    return path


def _normalize_technosphere_entries(entries):
    normalized = []
    for row in entries:
        if isinstance(row, dict):
            normalized.append(
                (
                    str(row["name"]),
                    str(row["reference_product"]),
                    str(row["location"]),
                    str(row["unit"]),
                )
            )
            continue
        normalized.append(tuple(row))
    return normalized


def _normalize_biosphere_entries(entries):
    normalized = []
    for row in entries:
        if isinstance(row, dict):
            normalized.append(
                (
                    str(row["name"]),
                    tuple(str(item) for item in row["categories"]),
                    str(row["unit"]),
                )
            )
            continue
        name, categories, unit = row
        normalized.append((str(name), tuple(str(item) for item in categories), str(unit)))
    return normalized


def write_catalog(tmp_path, *, family, version, system_model, technosphere, biosphere):
    directory = tmp_path / "reference_catalogs"
    directory.mkdir(exist_ok=True)
    normalized_technosphere = _normalize_technosphere_entries(technosphere)
    normalized_biosphere = _normalize_biosphere_entries(biosphere)
    path = directory / f"{family}__{version}__{system_model}.json"
    path.write_text(
        json.dumps(
            {
                "profile": {
                    "family": family,
                    "version": version,
                    "system_model": system_model,
                },
                "technosphere": [
                    {
                        "name": name,
                        "reference_product": reference_product,
                        "location": location,
                        "unit": unit,
                    }
                    for name, reference_product, location, unit in normalized_technosphere
                ],
                "biosphere": [
                    {
                        "name": name,
                        "categories": list(categories),
                        "unit": unit,
                    }
                    for name, categories, unit in normalized_biosphere
                ],
            }
        ),
        encoding="utf-8",
    )
    return directory


def test_infer_source_format_supports_xlsx_and_csv():
    assert infer_source_format("inventory.xlsx") == SOURCE_FORMAT_BRIGHTWAY_EXCEL
    assert infer_source_format("inventory.tsv") == SOURCE_FORMAT_BRIGHTWAY_TSV
    assert infer_source_format("inventory.csv") == SOURCE_FORMAT_SIMAPRO_CSV


def test_analyze_brightway_excel_returns_candidate_summaries(tmp_path):
    workbook = make_brightway_excel(tmp_path, [minimal_activity()])

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cut-off",
        ),
    )

    assert result.detected_software == "brightway"
    assert result.detected_format == SOURCE_FORMAT_BRIGHTWAY_EXCEL
    assert result.source_profile.family == "ecoinvent"
    assert result.source_profile.system_model == "cutoff"
    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert result.candidates[0].name == "test process"
    assert result.candidates[0].reference_product == "test product"
    assert result.candidates[0].description_hint == ""
    assert result.candidates[0].source_hint == ""
    assert result.candidates[0].issues == []


def test_analyze_brightway_excel_extracts_description_and_source_hints(tmp_path):
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                comment="Dataset comment",
                source="Journal article",
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert len(result.candidates) == 1
    assert result.candidates[0].description_hint == "Dataset comment"
    assert result.candidates[0].source_hint == "Journal article"


def test_analyze_brightway_excel_extracts_trailing_source_from_comment(tmp_path):
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                comment=(
                    "This dataset models a foreground process. "
                    "Source: Foteinis et al. (2023). https://doi.org/example"
                )
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert len(result.candidates) == 1
    assert result.candidates[0].description_hint == "This dataset models a foreground process."
    assert result.candidates[0].source_hint == "Foteinis et al. (2023). https://doi.org/example"


def test_analyze_brightway_excel_keeps_comment_when_source_marker_is_not_trailing(tmp_path):
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                comment=("Source: internal screening assumptions. " "This dataset models a foreground process.")
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert len(result.candidates) == 1
    assert (
        result.candidates[0].description_hint
        == "Source: internal screening assumptions. This dataset models a foreground process."
    )
    assert result.candidates[0].source_hint == ""


def test_analyze_brightway_excel_ignores_missing_simapro_category(tmp_path):
    invalid_activity = minimal_activity()
    del invalid_activity["exchanges"][0]["simapro category"]
    workbook = make_brightway_excel(tmp_path, [invalid_activity])

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert result.candidates[0].issues == []


def test_analyze_brightway_excel_surfaces_non_blocking_validation_warnings(tmp_path):
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "biosphere",
                        "name": "Water, river",
                        "categories": ("natural resource", "in water"),
                        "unit": "cubic meter",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert len(result.candidates[0].issues) == 1
    assert result.candidates[0].issues[0].severity == "warning"
    assert result.candidates[0].issues[0].code == "inventory_validation_warning"
    assert result.candidates[0].issues[0].path == "activity[0]"
    assert "no water release flows were found" in result.candidates[0].issues[0].message


def test_analyze_brightway_csv_returns_candidate_summaries(tmp_path):
    csv_path = make_brightway_delimited(tmp_path, [minimal_activity()], suffix=".csv")

    result = analyze_inventory(
        path=csv_path,
        source_format=SOURCE_FORMAT_BRIGHTWAY_CSV,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cut-off",
        ),
    )

    assert result.detected_software == "brightway"
    assert result.detected_format == SOURCE_FORMAT_BRIGHTWAY_CSV
    assert result.source_profile.system_model == "cutoff"
    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert result.candidates[0].name == "test process"
    assert result.candidates[0].issues == []


def test_analyze_brightway_tsv_returns_candidate_summaries(tmp_path):
    tsv_path = make_brightway_delimited(tmp_path, [minimal_activity()], delimiter="\t", suffix=".tsv")

    result = analyze_inventory(
        path=tsv_path,
        source_format=SOURCE_FORMAT_BRIGHTWAY_TSV,
    )

    assert result.detected_software == "brightway"
    assert result.detected_format == SOURCE_FORMAT_BRIGHTWAY_TSV
    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert result.candidates[0].name == "test process"


def test_analyze_brightway_excel_infers_background_profile_from_catalogs(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "market for steel",
                "reference_product": "steel",
                "location": "GLO",
                "unit": "kilogram",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for steel",
                        "reference product": "steel",
                        "location": "GLO",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.source_profile == BackgroundProfile(
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
    )
    assert any(issue.code == "background_profile_inferred" for issue in result.file_issues)


def test_analyze_brightway_excel_prefers_latest_cutoff_when_profile_matches_are_tied(tmp_path, monkeypatch):
    directory = tmp_path / "reference_catalogs"
    directory.mkdir(exist_ok=True)
    for version in ("3.9", "3.10"):
        for system_model in ("cutoff", "consequential"):
            write_catalog(
                tmp_path,
                family="ecoinvent",
                version=version,
                system_model=system_model,
                technosphere=[
                    {
                        "name": "market for steel",
                        "reference_product": "steel",
                        "location": "GLO",
                        "unit": "kilogram",
                    }
                ],
                biosphere=[],
            )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for steel",
                        "reference product": "steel",
                        "location": "GLO",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.source_profile == BackgroundProfile(
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
    )
    assert len(result.file_issues) == 1
    assert result.file_issues[0].severity == "warning"
    assert result.file_issues[0].code == "background_profile_assumed"
    assert "selected ecoinvent 3.10 cutoff" in result.file_issues[0].message
    assert "override" in result.file_issues[0].suggested_fix


def test_analyze_brightway_excel_flags_missing_background_targets(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "missing market",
                        "reference product": "missing product",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    },
                    {
                        "type": "biosphere",
                        "name": "mystery flow",
                        "categories": ("air", "urban air close to ground"),
                        "unit": "kilogram",
                        "amount": 1.0,
                    },
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    issue_codes = [issue.code for issue in result.candidates[0].issues]
    assert "unknown_technosphere_target" in issue_codes
    assert "unknown_biosphere_flow" in issue_codes


def test_analyze_brightway_excel_groups_unknown_technosphere_targets_per_candidate(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "missing market a",
                        "reference product": "missing product a",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    },
                    {
                        "type": "technosphere",
                        "name": "missing market b",
                        "reference product": "missing product b",
                        "location": "RER",
                        "unit": "megajoule",
                        "amount": 3.0,
                    },
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    technosphere_issues = [
        issue for issue in result.candidates[0].issues if issue.code == "unknown_technosphere_target"
    ]
    assert len(technosphere_issues) == 1
    assert "missing market a | missing product a | CH | kilogram" in technosphere_issues[0].message
    assert "missing market b | missing product b | RER | megajoule" in technosphere_issues[0].message
    assert technosphere_issues[0].suggested_fix


def test_analyze_brightway_excel_accepts_additional_foreground_targets(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "foreground provider",
                        "reference product": "foreground product",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    unresolved = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )
    assert any(issue.code == "unknown_technosphere_target" for issue in unresolved.candidates[0].issues)

    resolved = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
        additional_foreground_targets=[("foreground provider", "foreground product", "CH", "kilogram")],
    )

    assert all(issue.code != "unknown_technosphere_target" for issue in resolved.candidates[0].issues)


def test_analyze_brightway_excel_fills_missing_foreground_reference_product(tmp_path):
    provider = minimal_activity(
        name="foreground provider",
        **{"reference product": "foreground product"},
    )
    provider["exchanges"][0]["name"] = "foreground provider"
    provider["exchanges"][0]["reference product"] = "foreground product"

    consumer = minimal_activity(
        name="foreground consumer",
        **{"reference product": "consumer product"},
        extra_exchanges=[
            {
                "type": "technosphere",
                "name": "foreground provider",
                "reference product": "",
                "location": "GLO",
                "unit": "kilogram",
                "amount": 2.0,
            }
        ],
    )
    consumer["exchanges"][0]["name"] = "foreground consumer"
    consumer["exchanges"][0]["reference product"] = "consumer product"

    workbook = make_brightway_excel(
        tmp_path,
        [provider, consumer],
    )

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.candidates[1].issues == []
    assert result.inventory_data[1]["exchanges"][1]["reference product"] == "foreground product"


def test_analyze_brightway_excel_fills_missing_foreground_reference_product_canonically(tmp_path):
    provider = minimal_activity(
        name="Foreground provider",
        **{"reference product": "Foreground product"},
    )
    provider["location"] = "CH"
    provider["exchanges"][0]["name"] = "Foreground provider"
    provider["exchanges"][0]["reference product"] = "Foreground product"
    provider["exchanges"][0]["location"] = "CH"

    consumer = minimal_activity(
        name="foreground consumer",
        **{"reference product": "consumer product"},
        extra_exchanges=[
            {
                "type": "technosphere",
                "name": "  foreground   provider  ",
                "reference product": "",
                "location": "ch",
                "unit": "Kilogram",
                "amount": 2.0,
            }
        ],
    )
    consumer["exchanges"][0]["name"] = "foreground consumer"
    consumer["exchanges"][0]["reference product"] = "consumer product"

    workbook = make_brightway_excel(tmp_path, [provider, consumer])

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.candidates[1].issues == []
    assert result.inventory_data[1]["exchanges"][1]["name"] == "Foreground provider"
    assert result.inventory_data[1]["exchanges"][1]["reference product"] == "Foreground product"
    assert result.inventory_data[1]["exchanges"][1]["location"] == "CH"
    assert result.inventory_data[1]["exchanges"][1]["unit"] == "kilogram"


def test_analyze_brightway_excel_accepts_canonical_additional_foreground_targets(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "  foreground provider ",
                        "reference product": " foreground product ",
                        "location": "ch",
                        "unit": "Kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    resolved = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
        additional_foreground_targets=[("Foreground provider", "Foreground product", "CH", "kilogram")],
    )

    assert all(issue.code != "unknown_technosphere_target" for issue in resolved.candidates[0].issues)
    assert resolved.inventory_data[0]["exchanges"][1]["name"] == "Foreground provider"
    assert resolved.inventory_data[0]["exchanges"][1]["reference product"] == "Foreground product"
    assert resolved.inventory_data[0]["exchanges"][1]["location"] == "CH"
    assert resolved.inventory_data[0]["exchanges"][1]["unit"] == "kilogram"


def test_analyze_brightway_excel_accepts_canonical_background_catalog_targets(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "market for palladium",
                "reference_product": "palladium",
                "location": "GLO",
                "unit": "kilogram",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "  Market   for Palladium ",
                        "reference product": "Palladium",
                        "location": "glo",
                        "unit": "Kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert all(issue.code != "unknown_technosphere_target" for issue in result.candidates[0].issues)
    assert result.inventory_data[0]["exchanges"][1]["name"] == "market for palladium"
    assert result.inventory_data[0]["exchanges"][1]["reference product"] == "palladium"
    assert result.inventory_data[0]["exchanges"][1]["location"] == "GLO"
    assert result.inventory_data[0]["exchanges"][1]["unit"] == "kilogram"


def test_analyze_brightway_excel_accepts_technosphere_unit_alias_against_catalog(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "market for sowing",
                "reference_product": "sowing",
                "location": "GLO",
                "unit": "hectare",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for sowing",
                        "reference product": "sowing",
                        "location": "GLO",
                        "unit": "ha",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert result.candidates[0].issues == []
    assert result.inventory_data[0]["exchanges"][1]["unit"] == "hectare"


def test_analyze_brightway_excel_promotes_legacy_product_field_on_technosphere_exchange(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "assembly operation, for lorry",
                "reference_product": "assembly operation, for lorry",
                "location": "RER",
                "unit": "kilogram",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "assembly operation, for lorry",
                        "product": "assembly operation, for lorry",
                        "location": "RER",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert result.candidates[0].issues == []
    assert result.inventory_data[0]["exchanges"][1]["reference product"] == "assembly operation, for lorry"


def test_analyze_brightway_excel_accepts_person_kilometer_unit_alias_against_catalog(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "transport, tram",
                "reference_product": "transport, tram",
                "location": "CH",
                "unit": "person-kilometer",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "transport, tram",
                        "reference product": "transport, tram",
                        "location": "CH",
                        "unit": "person kilometer",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert result.candidates[0].issues == []
    assert result.inventory_data[0]["exchanges"][1]["unit"] == "person-kilometer"


def test_analyze_brightway_excel_fills_missing_catalog_reference_product(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[
            {
                "name": "market for steel",
                "reference_product": "steel",
                "location": "CH",
                "unit": "kilogram",
            }
        ],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for steel",
                        "reference product": "",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert result.candidates[0].issues == []
    assert result.inventory_data[0]["exchanges"][1]["reference product"] == "steel"


def test_analyze_brightway_excel_fills_missing_production_reference_product(tmp_path):
    activity = minimal_activity(
        name="foreground provider",
        **{"reference product": "foreground product"},
    )
    activity["exchanges"][0]["name"] = "foreground provider"
    activity["exchanges"][0]["reference product"] = ""

    workbook = make_brightway_excel(tmp_path, [activity])

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.candidates[0].issues == []
    assert result.inventory_data[0]["exchanges"][0]["reference product"] == "foreground product"


def test_analyze_brightway_excel_normalizes_biosphere_aliases_before_link_check(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[
            {
                "name": "Particulate Matter, < 2.5 um",
                "categories": ["air", "urban air close to ground"],
                "unit": "kilogram",
            }
        ],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "biosphere",
                        "name": "Particulates, < 2.5 um",
                        "categories": ("air", "urban air close to ground"),
                        "unit": "kilogram",
                        "amount": 0.1,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.10",
            system_model="cutoff",
        ),
    )

    assert all(issue.code != "unknown_biosphere_flow" for issue in result.candidates[0].issues)
    assert result.inventory_data[0]["exchanges"][1]["name"] == "Particulate Matter, < 2.5 um"


def test_analyze_brightway_excel_maps_biosphere_alias_to_selected_catalog_name(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.8",
        system_model="cutoff",
        technosphere=[],
        biosphere=[
            {
                "name": "Selenium",
                "categories": ["air", "non-urban air or from high stacks"],
                "unit": "kilogram",
            }
        ],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "biosphere",
                        "name": "Selenium IV",
                        "categories": ("air", "non-urban air or from high stacks"),
                        "unit": "kilogram",
                        "amount": 0.1,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.8",
            system_model="cutoff",
        ),
    )

    assert all(issue.code != "unknown_biosphere_flow" for issue in result.candidates[0].issues)
    assert result.inventory_data[0]["exchanges"][1]["name"] == "Selenium"


def test_analyze_brightway_excel_applies_supplemental_biosphere_aliases(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.12",
        system_model="cutoff",
        technosphere=[],
        biosphere=[
            {
                "name": "Propylene",
                "categories": ["air", "urban air close to ground"],
                "unit": "kilogram",
            }
        ],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "biosphere",
                        "name": "Propene",
                        "categories": ("air", "urban air close to ground"),
                        "unit": "kilogram",
                        "amount": 0.1,
                    }
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.12",
            system_model="cutoff",
        ),
    )

    assert all(issue.code != "unknown_biosphere_flow" for issue in result.candidates[0].issues)
    assert result.inventory_data[0]["exchanges"][1]["name"] == "Propylene"


def test_analyze_brightway_excel_accepts_uvek_catalog_with_external_biosphere_reference(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="uvek",
        version="2025",
        system_model="cutoff",
        technosphere=[
            {
                "name": "market for steel",
                "reference_product": "steel",
                "location": "CH",
                "unit": "kilogram",
            }
        ],
        biosphere=[
            {
                "name": "Carbon dioxide, fossil",
                "categories": ["air"],
                "unit": "kilogram",
            }
        ],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for steel",
                        "reference product": "steel",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    },
                    {
                        "type": "biosphere",
                        "name": "Carbon dioxide, fossil",
                        "categories": ("air",),
                        "unit": "kilogram",
                        "amount": 1.0,
                    },
                ]
            )
        ],
    )

    result = analyze_inventory(
        path=workbook,
        source_profile=BackgroundProfile(
            family="uvek",
            version="2025",
            system_model="cutoff",
        ),
    )

    assert result.file_issues == []
    assert result.candidates[0].issues == []


def test_analyze_brightway_excel_keeps_structural_validation_errors_blocking(tmp_path):
    invalid_activity = minimal_activity()
    invalid_activity["exchanges"][0]["unit"] = "made-up unit"
    workbook = make_brightway_excel(tmp_path, [invalid_activity])

    result = analyze_inventory(path=workbook, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert len(result.candidates[0].issues) == 1
    assert result.candidates[0].issues[0].severity == "error"
    assert result.candidates[0].issues[0].code == "inventory_validation_error"
    assert "unknown exchange unit" in result.candidates[0].issues[0].message


def test_validate_inventory_raises_for_unknown_background_targets(tmp_path, monkeypatch):
    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.10",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "missing market",
                        "reference product": "missing product",
                        "location": "CH",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    try:
        validate_inventory(
            path=workbook,
            source_profile=BackgroundProfile(
                family="ecoinvent",
                version="3.10",
                system_model="cutoff",
            ),
        )
    except InventoryValidationError as exc:
        assert exc.result.candidates[0].issues[0].code == "unknown_technosphere_target"
        assert "Technosphere exchanges do not match" in str(exc)
    else:
        raise AssertionError("InventoryValidationError was not raised")


def test_validate_inventory_raises_when_background_catalog_is_missing(tmp_path, monkeypatch):
    directory = tmp_path / "reference_catalogs"
    directory.mkdir()
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    workbook = make_brightway_excel(
        tmp_path,
        [
            minimal_activity(
                extra_exchanges=[
                    {
                        "type": "technosphere",
                        "name": "market for steel",
                        "reference product": "steel",
                        "location": "GLO",
                        "unit": "kilogram",
                        "amount": 2.0,
                    }
                ]
            )
        ],
    )

    try:
        validate_inventory(
            path=workbook,
            source_profile=BackgroundProfile(
                family="ecoinvent",
                version="3.5",
                system_model="cutoff",
            ),
        )
    except InventoryValidationError as exc:
        assert exc.result.file_issues[0].code == "background_catalog_missing"
        assert "No local reference catalog is available" in str(exc)
    else:
        raise AssertionError("InventoryValidationError was not raised")


def test_analyze_simapro_csv_returns_candidate_summaries(tmp_path):
    filepath = make_simapro_csv(tmp_path, [minimal_activity()])

    result = analyze_inventory(
        path=filepath,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.9",
            system_model="cutoff",
        ),
    )

    assert result.detected_software == "simapro"
    assert result.detected_format == SOURCE_FORMAT_SIMAPRO_CSV
    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert result.candidates[0].name == "test process"
    assert result.candidates[0].reference_product == "test product"
    assert result.candidates[0].location == "GLO"


def test_analyze_simapro_csv_surfaces_validation_warnings_from_converted_inventory(tmp_path, monkeypatch):
    from brightpath.analysis import analyzer as analysis_analyzer

    filepath = tmp_path / "inventory.csv"
    filepath.write_text("fake simapro content", encoding="utf-8")
    inventory_data = [
        minimal_activity(
            extra_exchanges=[
                {
                    "type": "biosphere",
                    "name": "Water, river",
                    "categories": ("natural resource", "in water"),
                    "unit": "cubic meter",
                    "amount": 2.0,
                }
            ]
        )
    ]

    class FakeSimaProInventory:
        data = inventory_data

        @classmethod
        def from_csv(cls, *args, **kwargs):
            return cls()

        def validate(self, **kwargs):
            return type("Report", (), {"issues": []})()

    monkeypatch.setattr(analysis_analyzer, "SimaProInventory", FakeSimaProInventory)

    result = analyze_inventory(path=filepath, source_format=SOURCE_FORMAT_SIMAPRO_CSV)

    assert result.file_issues == []
    assert len(result.candidates) == 1
    assert any(
        issue.code == "inventory_validation_warning" and "no water release flows were found" in issue.message
        for issue in result.candidates[0].issues
    )


def test_analyze_simapro_csv_validates_background_links_from_converted_inventory(tmp_path, monkeypatch):
    from brightpath.analysis import analyzer as analysis_analyzer

    directory = write_catalog(
        tmp_path,
        family="ecoinvent",
        version="3.9",
        system_model="cutoff",
        technosphere=[],
        biosphere=[],
    )
    monkeypatch.setenv("BRIGHTPATH_REFERENCE_DIR", str(directory))
    filepath = tmp_path / "inventory.csv"
    filepath.write_text("fake simapro content", encoding="utf-8")
    inventory_data = [
        minimal_activity(
            extra_exchanges=[
                {
                    "type": "technosphere",
                    "name": "missing market",
                    "reference product": "missing product",
                    "location": "CH",
                    "unit": "kilogram",
                    "amount": 2.0,
                }
            ]
        )
    ]

    class FakeSimaProInventory:
        data = inventory_data

        @classmethod
        def from_csv(cls, *args, **kwargs):
            return cls()

        def validate(self, **kwargs):
            return type("Report", (), {"issues": []})()

    monkeypatch.setattr(analysis_analyzer, "SimaProInventory", FakeSimaProInventory)

    result = analyze_inventory(
        path=filepath,
        source_format=SOURCE_FORMAT_SIMAPRO_CSV,
        source_profile=BackgroundProfile(
            family="ecoinvent",
            version="3.9",
            system_model="cutoff",
        ),
    )

    assert len(result.candidates) == 1
    assert any(issue.code == "unknown_technosphere_target" for issue in result.candidates[0].issues)


def test_analyze_simapro_csv_attaches_duplicate_identity_errors(tmp_path):
    filepath = make_simapro_csv(tmp_path, [minimal_activity(), minimal_activity()], filename="duplicates.csv")

    result = analyze_inventory(path=filepath, source_format=SOURCE_FORMAT_SIMAPRO_CSV)

    assert result.file_issues == []
    assert len(result.candidates) == 2
    assert all(candidate.issues for candidate in result.candidates)
    assert all(candidate.issues[0].code == "duplicate_dataset_identity" for candidate in result.candidates)


def test_analyze_inventory_reports_missing_file_as_file_error(tmp_path):
    missing = tmp_path / "missing.xlsx"

    result = analyze_inventory(path=missing, source_format=SOURCE_FORMAT_BRIGHTWAY_EXCEL)

    assert len(result.file_issues) == 1
    assert result.file_issues[0].code == "brightway_excel_parse_failed"
    assert "could not be found" in result.file_issues[0].message
